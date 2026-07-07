#!/usr/bin/env python3
"""
Full inventory extraction from /tmp/inventory_spec.xlsx
Parses all content tabs, vehicle equipment, שו"ב distribution, and דוח צ.
Generates SQL with proper quantity handling.
"""

import openpyxl
import re
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

EXCEL_PATH = '/tmp/inventory_spec.xlsx'
OUTPUT_SQL = '/private/tmp/inventory-app/update_db_v2.sql'
ALTER_SQL = '/private/tmp/inventory-app/alter_table.sql'

# Content tabs: (sheet_name, company_override)
CONTENT_TABS = [
    ('תכולה פל א', None),
    ('תכולה פל ב', None),
    ('תכולה פל ג', None),
    ('טופס 1 תכולה  - מסייעת ', None),
    ('טופס תכולה  מפג"ד', None),
]

# Serial number columns (header name -> True means this col IS the serial for the PREVIOUS product)
SERIAL_COLUMNS = {
    "צ' מכשיר",
    "צ' מכשיר טל 88",
    "צ' אול\"ר",
}

# Map product columns to their serial columns (by header name)
PRODUCT_TO_SERIAL = {
    'מכשיר 710': "צ' מכשיר",
    'מכשיר 711': "צ' מכשיר",
    'טל 88': "צ' מכשיר טל 88",
    'אול"ר': "צ' אול\"ר",
}


def escape_sql(val):
    """Escape single quotes for SQL and remove newlines."""
    if val is None:
        return ''
    s = str(val).strip()
    s = s.replace('\n', ' ').replace('\r', ' ')
    s = s.replace("'", "''")
    return s


def clean_value(val):
    """Clean a cell value - return None if empty or #REF!."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '#REF!', 'None', 'none'):
        return None
    return s


def to_int_qty(val):
    """Convert a value to integer quantity. Returns 0 if not numeric."""
    if val is None:
        return 0
    s = str(val).strip()
    if s in ('', '#REF!', 'None'):
        return 0
    try:
        n = float(s)
        return int(n)
    except (ValueError, TypeError):
        return 0


def is_total_row(role):
    """Check if this is a totals row to skip."""
    if not role:
        return True
    role_str = str(role).strip()
    if not role_str:
        return True
    if 'סה"כ' in role_str or 'סה״כ' in role_str:
        return True
    return False


def parse_content_tab(ws, sheet_name):
    """Parse a standard content tab (תכולה)."""
    items = []

    # Read headers from row 5
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(5, c).value
        if v:
            h = str(v).strip().replace('\n', ' ').replace('\r', ' ')
            headers[c] = h

    # Identify product columns (not serial, not metadata)
    metadata_cols = {3, 4, 5}  # location, crate, role (1-indexed: col 3,4,5)
    serial_col_indices = set()
    product_cols = []

    # Build serial column index map
    serial_header_to_col = {}
    for c, h in headers.items():
        if h in SERIAL_COLUMNS:
            serial_col_indices.add(c)
            serial_header_to_col[h] = c

    # Build product column list (skip metadata, serial, #REF! columns)
    for c, h in headers.items():
        if c in metadata_cols or c <= 2:
            continue
        if c in serial_col_indices:
            continue
        if h == '#REF!':
            continue
        product_cols.append((c, h))

    # Build product-to-serial-col mapping
    product_serial_map = {}
    for prod_name, serial_header in PRODUCT_TO_SERIAL.items():
        if serial_header in serial_header_to_col:
            product_serial_map[prod_name] = serial_header_to_col[serial_header]

    # Parse data rows starting from row 6
    current_company = None

    for r in range(6, ws.max_row + 1):
        # Company (col 2) - propagate from merged cells
        company_val = ws.cell(r, 2).value
        if company_val and str(company_val).strip():
            current_company = str(company_val).strip()

        # Role (col 5)
        role = clean_value(ws.cell(r, 5).value)
        if is_total_row(role):
            # If we hit a total row, skip it
            if role and 'סה"כ' in str(role):
                continue
            # Empty role - check if there's any data in the row
            if not role:
                has_data = False
                for c, h in product_cols:
                    if clean_value(ws.cell(r, c).value):
                        has_data = True
                        break
                if not has_data:
                    continue
                else:
                    role = ''  # Row has data but no role
            else:
                continue

        location = clean_value(ws.cell(r, 3).value) or ''
        crate = clean_value(ws.cell(r, 4).value) or ''

        if not current_company:
            current_company = ''

        # Extract products
        for col_idx, prod_name in product_cols:
            qty = to_int_qty(ws.cell(r, col_idx).value)
            if qty <= 0:
                continue

            # Check for serial number
            serial = ''
            if prod_name in product_serial_map:
                serial_col = product_serial_map[prod_name]
                serial_val = clean_value(ws.cell(r, serial_col).value)
                if serial_val:
                    # Clean serial - convert float strings to int
                    try:
                        serial = str(int(float(serial_val)))
                    except (ValueError, TypeError):
                        serial = serial_val

            items.append({
                'company': current_company,
                'role': role,
                'product': prod_name,
                'serial': serial,
                'location': location,
                'crate': crate,
                'quantity': qty,
            })

    return items


def parse_vehicle_tab(ws):
    """Parse טופס זיווד רכבים tab."""
    items = []

    # Headers in row 4
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(4, c).value
        if v:
            h = str(v).strip().replace('\n', ' ')
            headers[c] = h

    # Serial columns for vehicles (typically paired: qty col followed by serial col)
    # Pattern: product col has # (quantity), next col has צ' (serial)
    serial_cols = set()
    product_to_serial = {}

    for c, h in headers.items():
        if h.startswith("צ'") or h.startswith('צ ') or "צ'" in h:
            serial_cols.add(c)

    # Map products to their serial columns based on adjacency
    # Known pairings from the header structure:
    vehicle_serial_pairs = {
        13: 14,   # מגן מכלול -> צ' מגן מכלול
        14: None, # צ' מגן מכלול (serial itself)
        15: None, # צ' מדיה נתיקה
        16: 17,   # נר לילה -> צ' נר לילה
        21: 22,   # שן בינה -> צ' שן בינה
        23: None, # צ' מדיה נתיקה
        24: 25,   # אלעד ירוק -> צ' אלעד ירוק
        26: None, # צ' מדיה נתיקה
        33: 34,   # טל 88 -> צ' טל 88
        35: 36,   # מ"ק 376 -> צ' חד-פס
        47: 48,   # מב"ן -> צ' מב"ן
        49: 50,   # RPT -> צ' RPT
        51: 52,   # CF33 -> צ' CF33
        53: 54,   # CF55 -> צ' CF55
        55: 56,   # מפל"ז -> צ' מפל"ז
        59: 60,   # דיסקית -> צ' דיסקית
        62: 63,   # גנרטור -> צ' גנרטור
    }

    # Product columns (not serial, not metadata)
    metadata_cols = {1, 2, 3, 4, 5, 6, 7, 8}  # #, מסגרת, צ' רכב, סוג, מוכנות, blank, הערות, דיגום

    for r in range(5, ws.max_row + 1):
        vehicle_type = clean_value(ws.cell(r, 5).value)
        if not vehicle_type:
            continue
        if 'סה"כ' in str(vehicle_type):
            continue

        company = clean_value(ws.cell(r, 3).value) or ''  # מסגרת משנה = company/unit
        vehicle_serial = clean_value(ws.cell(r, 4).value) or ''  # צ' רכב
        if vehicle_serial:
            try:
                vehicle_serial = str(int(float(vehicle_serial)))
            except (ValueError, TypeError):
                pass

        role = f"{vehicle_type} ({vehicle_serial})" if vehicle_serial else vehicle_type

        # The vehicle itself is an item
        items.append({
            'company': company,
            'role': role,
            'product': 'רכב - ' + vehicle_type,
            'serial': vehicle_serial,
            'location': 'רכב',
            'crate': '',
            'quantity': 1,
        })

        # Parse equipment columns
        for c in range(9, ws.max_column + 1):
            if c in serial_cols or c in metadata_cols:
                continue
            h = headers.get(c, '')
            if not h or h == '#REF!':
                continue
            if h.startswith("צ'") or h.startswith('צ '):
                continue

            val = ws.cell(r, c).value
            qty = to_int_qty(val)

            # Column 9 (התקן מוסע) is a text description, not quantity
            if c == 9:
                continue

            if qty <= 0:
                continue

            # Check for serial
            serial = ''
            serial_col = vehicle_serial_pairs.get(c)
            if serial_col:
                serial_val = clean_value(ws.cell(r, serial_col).value)
                if serial_val:
                    try:
                        serial = str(int(float(serial_val)))
                    except (ValueError, TypeError):
                        serial = serial_val

            items.append({
                'company': company,
                'role': role,
                'product': h,
                'serial': serial,
                'location': 'רכב',
                'crate': '',
                'quantity': qty,
            })

    return items


def parse_shub_tab(ws):
    """Parse פיזור שו"ב tab - 3 sections."""
    items = []

    # Section 1: Portable שו"ב (rows 5-18)
    # Headers in row 4: פלוגה, בעלי תפקיד, אוק, מב"ן, מפל"ז, קוד מפלז, G2, צרעה, סל"צ - שוב, סים אדום, מטען נייד לG2, קורא כרטיסים, תיק לG2
    section1_headers = {}
    for c in range(1, 17):
        v = ws.cell(4, c).value
        if v:
            section1_headers[c] = str(v).strip()

    # Serial-type columns in section 1 (items with unique IDs)
    serial_products_s1 = {4: 'מב"ן', 5: 'מפל"ז', 6: 'קוד מפלז', 7: 'G2', 8: 'צרעה', 9: 'סל"צ - שוב', 10: 'סים אדום'}
    qty_products_s1 = {11: 'מטען נייד לG2', 12: 'קורא כרטיסים', 13: 'תיק לG2'}

    for r in range(5, 19):
        company = clean_value(ws.cell(r, 1).value) or ''
        role = clean_value(ws.cell(r, 2).value) or ''
        if not role or is_total_row(role):
            continue

        # Serial products
        for c, prod in serial_products_s1.items():
            val = clean_value(ws.cell(r, c).value)
            if val:
                items.append({
                    'company': company,
                    'role': role,
                    'product': prod,
                    'serial': val,
                    'location': 'שו"ב נייד',
                    'crate': '',
                    'quantity': 1,
                })

        # Qty products
        for c, prod in qty_products_s1.items():
            qty = to_int_qty(ws.cell(r, c).value)
            if qty > 0:
                items.append({
                    'company': company,
                    'role': role,
                    'product': prod,
                    'serial': '',
                    'location': 'שו"ב נייד',
                    'crate': '',
                    'quantity': qty,
                })

    # Section 2: שו"ב רכובים (rows 22-42)
    # Headers in row 21: פלוגה, בעלי תפקיד, אוק, מב"ן, מפל"ז, קוד מפלז, CF33, CF55, RPT, ברק, מדיה לברק, שן בינה, מדיה לשן בינה, דיסקית, קורא כרטיסים
    serial_products_s2 = {4: 'מב"ן', 5: 'מפל"ז', 6: 'קוד מפלז', 7: 'CF33', 8: 'CF55', 9: 'RPT', 10: 'ברק', 11: 'מדיה לברק', 12: 'שן בינה', 13: 'מדיה לשן בינה', 14: 'דיסקית'}
    qty_products_s2 = {15: 'קורא כרטיסים'}

    for r in range(22, 43):
        company = clean_value(ws.cell(r, 1).value) or ''
        role = clean_value(ws.cell(r, 2).value) or ''
        if not role or is_total_row(role) or role == 'בעלי תפקיד':
            continue

        for c, prod in serial_products_s2.items():
            val = clean_value(ws.cell(r, c).value)
            if val:
                items.append({
                    'company': company,
                    'role': role,
                    'product': f'{prod} (רכוב)',
                    'serial': val,
                    'location': 'שו"ב רכוב',
                    'crate': '',
                    'quantity': 1,
                })

        for c, prod in qty_products_s2.items():
            qty = to_int_qty(ws.cell(r, c).value)
            if qty > 0:
                items.append({
                    'company': company,
                    'role': role,
                    'product': f'{prod} (רכוב)',
                    'serial': '',
                    'location': 'שו"ב רכוב',
                    'crate': '',
                    'quantity': qty,
                })

    # Section 3: סלצי"ם ואולרי"ם רשתיים (rows 46-64)
    # Headers row 45: פלוגה, בעלי תפקיד, אוק, סל"צ, סים אדום, אול"ר רשתי
    serial_products_s3 = {4: 'סל"צ', 5: 'סים אדום', 6: 'אול"ר רשתי'}

    for r in range(46, 65):
        company = clean_value(ws.cell(r, 1).value) or ''
        role = clean_value(ws.cell(r, 2).value) or ''
        if not role or is_total_row(role) or role == 'בעלי תפקיד':
            continue

        for c, prod in serial_products_s3.items():
            val = clean_value(ws.cell(r, c).value)
            if val:
                items.append({
                    'company': company,
                    'role': role,
                    'product': prod,
                    'serial': val,
                    'location': 'שו"ב רשתי',
                    'crate': '',
                    'quantity': 1,
                })

    return items


def parse_doch_tz(ws):
    """Parse דוח צ - עדכני tab."""
    items = []

    # Headers row 3: פלוגה, מיקום, בעלי תפקיד, שם, מכשיר, צ, שיוך, date, הערות
    for r in range(4, ws.max_row + 1):
        company = clean_value(ws.cell(r, 1).value) or ''
        role = clean_value(ws.cell(r, 3).value) or ''
        product = clean_value(ws.cell(r, 5).value) or ''
        serial = clean_value(ws.cell(r, 6).value) or ''

        if not role or not product:
            continue
        if is_total_row(role):
            continue

        # Clean serial (often float like 102312496.0)
        if serial:
            try:
                serial = str(int(float(serial)))
            except (ValueError, TypeError):
                pass

        assignment = clean_value(ws.cell(r, 7).value) or ''  # שיוך
        notes = clean_value(ws.cell(r, 9).value) or ''
        location = f'דוח צ - {assignment}' if assignment else 'דוח צ'

        items.append({
            'company': company,
            'role': role,
            'product': product,
            'serial': serial,
            'location': location,
            'crate': '',
            'quantity': 1,
        })

    return items


def generate_sql(all_items):
    """Generate SQL INSERT statements."""
    lines = []
    lines.append("-- Auto-generated inventory SQL from parse_full.py")
    lines.append("-- Source: /tmp/inventory_spec.xlsx")
    lines.append(f"-- Total items: {len(all_items)}")
    lines.append("")
    lines.append("TRUNCATE TABLE equipment RESTART IDENTITY;")
    lines.append("")

    if not all_items:
        return '\n'.join(lines)

    lines.append("INSERT INTO equipment (company, role, product, serial, location, crate, quantity) VALUES")

    value_lines = []
    for item in all_items:
        company = escape_sql(item['company'])
        role = escape_sql(item['role'])
        product = escape_sql(item['product'])
        serial = escape_sql(item['serial'])
        location = escape_sql(item['location'])
        crate = escape_sql(item['crate'])
        quantity = item['quantity']

        value_lines.append(
            f"('{company}', '{role}', '{product}', '{serial}', '{location}', '{crate}', {quantity})"
        )

    # Join with commas
    for i, vl in enumerate(value_lines):
        if i < len(value_lines) - 1:
            lines.append(f"  {vl},")
        else:
            lines.append(f"  {vl};")

    return '\n'.join(lines)


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_items = []

    # 1. Parse content tabs
    for sheet_name, _ in CONTENT_TABS:
        print(f"Parsing: {sheet_name}")
        ws = wb[sheet_name]
        items = parse_content_tab(ws, sheet_name)
        print(f"  -> {len(items)} items")
        all_items.extend(items)

    # 2. Parse vehicle tab
    print("Parsing: טופס זיווד רכבים")
    ws = wb['טופס זיווד רכבים']
    items = parse_vehicle_tab(ws)
    print(f"  -> {len(items)} items")
    all_items.extend(items)

    # 3. Parse שו"ב distribution
    print('Parsing: פיזור שו"ב')
    ws = wb['פיזור שו"ב']
    items = parse_shub_tab(ws)
    print(f"  -> {len(items)} items")
    all_items.extend(items)

    # 4. Parse דוח צ - עדכני
    print("Parsing: דוח צ - עדכני")
    ws = wb['דוח צ - עדכני']
    items = parse_doch_tz(ws)
    print(f"  -> {len(items)} items")
    all_items.extend(items)

    print(f"\nTotal items extracted: {len(all_items)}")

    # Generate ALTER TABLE SQL
    alter_sql = "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS quantity INTEGER DEFAULT 1;\n"
    with open(ALTER_SQL, 'w', encoding='utf-8') as f:
        f.write(alter_sql)
    print(f"Written: {ALTER_SQL}")

    # Generate INSERT SQL
    sql = generate_sql(all_items)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"Written: {OUTPUT_SQL}")

    # Print summary by company
    from collections import Counter
    company_counts = Counter(item['company'] for item in all_items)
    print("\nItems by company:")
    for company, count in sorted(company_counts.items(), key=lambda x: -x[1]):
        print(f"  {company or '(empty)'}: {count}")

    # Print summary by source/location
    location_counts = Counter(item['location'] for item in all_items)
    print("\nItems by location/source:")
    for loc, count in sorted(location_counts.items(), key=lambda x: -x[1]):
        print(f"  {loc or '(empty)'}: {count}")


if __name__ == '__main__':
    main()
